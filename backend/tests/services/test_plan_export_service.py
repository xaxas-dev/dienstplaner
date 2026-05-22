"""Unit-Tests für plan_export_service.build_plan_xlsx."""
from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

import app.models  # noqa: F401 – alle Modelle registrieren
from app.models.doctor import Doctor
from app.models.plan import Plan, PlanStatus
from app.models.shift import Shift
from app.models.shift_type import ShiftType
from app.services.exceptions import PlanNotFoundError
from app.services.plan_export_service import build_plan_xlsx, make_filename_slug


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------


def _make_shift_type(db: Session, *, short_name: str, name: str, display_order: int = 0) -> ShiftType:
    st = ShiftType(
        short_name=short_name,
        name=name,
        display_order=display_order,
        applies_on_weekdays=True,
        applies_on_weekend=False,
        active=True,
    )
    db.add(st)
    db.flush()
    return st


def _make_doctor(db: Session, *, name: str, short_name: str) -> Doctor:
    from app.models.doctor import DoctorType

    doc = Doctor(
        name=name,
        short_name=short_name,
        doctor_type=DoctorType.INTERNAL,
        active=True,
    )
    db.add(doc)
    db.flush()
    return doc


def _make_plan(db: Session, *, name: str = "Testplan Juni") -> Plan:
    p = Plan(
        name=name,
        valid_from=date(2026, 6, 1),
        valid_to=date(2026, 6, 30),
        status=PlanStatus.DRAFT,
    )
    db.add(p)
    db.flush()
    return p


def _make_shift(
    db: Session,
    *,
    plan: Plan,
    shift_type: ShiftType,
    shift_date: date,
    doctor: Doctor | None = None,
    is_pinned: bool = False,
    notes: str | None = None,
) -> Shift:
    s = Shift(
        plan_id=plan.id,
        shift_type_id=shift_type.id,
        shift_date=shift_date,
        doctor_id=doctor.id if doctor else None,
        is_pinned=is_pinned,
        notes=notes,
    )
    db.add(s)
    db.flush()
    return s


def _load_wb(data: bytes):
    return load_workbook(BytesIO(data))


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_empty_plan_has_header_no_data_rows(db: Session) -> None:
    plan = _make_plan(db)
    result = build_plan_xlsx(db, plan.id)
    wb = _load_wb(result)
    ws = wb["Dienste"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1  # nur Header
    assert rows[0][0] == "Datum"
    assert rows[0][1] == "Wochentag"


def test_plan_with_shifts_produces_one_row_per_shift(db: Session) -> None:
    plan = _make_plan(db)
    st_v = _make_shift_type(db, short_name="V", name="Vortagsdienst", display_order=1)
    st_t = _make_shift_type(db, short_name="T", name="Tagesdienst", display_order=2)
    st_n = _make_shift_type(db, short_name="N", name="Nachtdienst", display_order=3)
    doc = _make_doctor(db, name="Max Mustermann", short_name="MM")

    _make_shift(db, plan=plan, shift_type=st_v, shift_date=date(2026, 6, 1), doctor=doc)
    _make_shift(db, plan=plan, shift_type=st_t, shift_date=date(2026, 6, 1))  # kein Doctor
    _make_shift(db, plan=plan, shift_type=st_n, shift_date=date(2026, 6, 2))

    result = build_plan_xlsx(db, plan.id)
    wb = _load_wb(result)
    ws = wb["Dienste"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 4  # 1 Header + 3 Datenzeilen


def test_shift_without_doctor_produces_empty_arzt_cells(db: Session) -> None:
    plan = _make_plan(db)
    st = _make_shift_type(db, short_name="V2", name="Vortagsdienst2", display_order=1)
    _make_shift(db, plan=plan, shift_type=st, shift_date=date(2026, 6, 3))

    result = build_plan_xlsx(db, plan.id)
    wb = _load_wb(result)
    ws = wb["Dienste"]
    row = list(ws.iter_rows(min_row=2, values_only=True))[0]
    # Arzt-Kürzel (Index 4) und Arzt (Index 5) leer
    assert row[4] is None or row[4] == ""
    assert row[5] is None or row[5] == ""


def test_rows_sorted_by_date_then_shifttype_order(db: Session) -> None:
    plan = _make_plan(db)
    st_a = _make_shift_type(db, short_name="SA", name="ShiftA", display_order=2)
    st_b = _make_shift_type(db, short_name="SB", name="ShiftB", display_order=1)

    # Einfüge-Reihenfolge bewusst umgekehrt
    _make_shift(db, plan=plan, shift_type=st_a, shift_date=date(2026, 6, 5))
    _make_shift(db, plan=plan, shift_type=st_b, shift_date=date(2026, 6, 5))
    _make_shift(db, plan=plan, shift_type=st_a, shift_date=date(2026, 6, 4))

    result = build_plan_xlsx(db, plan.id)
    wb = _load_wb(result)
    ws = wb["Dienste"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 3
    # Erste Zeile: 2026-06-04 (früheres Datum)
    assert rows[0][0] == "2026-06-04"
    # Zweite und dritte Zeile: 2026-06-05, geordnet nach display_order (SB=1 vor SA=2)
    assert rows[1][0] == "2026-06-05"
    assert rows[1][2] == "SB"
    assert rows[2][2] == "SA"


def test_pinned_shift_shows_ja(db: Session) -> None:
    plan = _make_plan(db)
    st = _make_shift_type(db, short_name="VP", name="VortagsPinned", display_order=1)
    _make_shift(db, plan=plan, shift_type=st, shift_date=date(2026, 6, 6), is_pinned=True)

    result = build_plan_xlsx(db, plan.id)
    wb = _load_wb(result)
    ws = wb["Dienste"]
    row = list(ws.iter_rows(min_row=2, values_only=True))[0]
    assert row[6] == "ja"  # Gepinnt-Spalte


def test_weekday_german_abbreviation(db: Session) -> None:
    # 2026-06-01 ist ein Montag
    plan = _make_plan(db)
    st = _make_shift_type(db, short_name="WD", name="WochentagTest", display_order=1)
    _make_shift(db, plan=plan, shift_type=st, shift_date=date(2026, 6, 1))

    result = build_plan_xlsx(db, plan.id)
    wb = _load_wb(result)
    ws = wb["Dienste"]
    row = list(ws.iter_rows(min_row=2, values_only=True))[0]
    assert row[1] == "Mo"


def test_unknown_plan_id_raises_plan_not_found_error(db: Session) -> None:
    with pytest.raises(PlanNotFoundError):
        build_plan_xlsx(db, 999999)


def test_bytes_roundtrip_opens_as_workbook(db: Session) -> None:
    plan = _make_plan(db)
    st = _make_shift_type(db, short_name="RT", name="Roundtrip", display_order=1)
    doc = _make_doctor(db, name="Dr. Test", short_name="DT")
    _make_shift(
        db,
        plan=plan,
        shift_type=st,
        shift_date=date(2026, 6, 10),
        doctor=doc,
        notes="Testnotiz",
    )

    result = build_plan_xlsx(db, plan.id)
    assert isinstance(result, bytes)
    assert len(result) > 0

    wb = _load_wb(result)
    assert "Dienste" in wb.sheetnames
    ws = wb["Dienste"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 2  # Header + 1 Datenzeile
    assert rows[1][0] == "2026-06-10"
    assert rows[1][4] == "DT"
    assert rows[1][5] == "Dr. Test"
    assert rows[1][7] == "Testnotiz"


# ---------------------------------------------------------------------------
# make_filename_slug
# ---------------------------------------------------------------------------


def test_slug_normal_name() -> None:
    assert make_filename_slug("Juni 2026", 1) == "Juni-2026.xlsx"


def test_slug_sonderzeichen_werden_entfernt() -> None:
    assert make_filename_slug("Plan/Juni:2026", 1) == "Plan-Juni-2026.xlsx"


def test_slug_leerer_name_fallback() -> None:
    assert make_filename_slug("", 42) == "plan-42.xlsx"
