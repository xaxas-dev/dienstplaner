"""Tests für den Import-Commit-Service (Phase C)."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from sqlalchemy.orm import Session

from app.models.department import Department
from app.models.doctor import Doctor
from app.models.employment_period import EmploymentPeriod
from app.models.rotation_assignment import RotationAssignment
from app.models.shift_type import ShiftType
from app.repositories import department_repository as dept_repo
from app.schemas.excel_import import CommitResolutions
from app.services import import_commit_service
from app.services.import_parse_service import parse_besetzungsplan

_FIXTURE = Path(__file__).parents[1] / "fixtures" / "Besetzungsplan_07_2026_V5.xlsx"


@pytest.fixture(autouse=True)
def seed_shift_types(db):
    """create_plan_with_shifts braucht mindestens einen aktiven ShiftType.
    In Produktion sind diese geseedet; im Test einen Werktags- und einen
    Wochenend-Typ anlegen, damit die Schichtgenerierung funktioniert.
    """
    db.add_all(
        [
            ShiftType(
                name="Vordergrund",
                short_name="V",
                applies_on_weekdays=True,
                applies_on_weekend=False,
            ),
            ShiftType(
                name="Tagdienst",
                short_name="T",
                applies_on_weekdays=False,
                applies_on_weekend=True,
            ),
        ]
    )
    db.commit()


@pytest.fixture
def file_bytes() -> bytes:
    return _FIXTURE.read_bytes()


@pytest.fixture
def parsed(file_bytes: bytes):
    return parse_besetzungsplan(file_bytes)


def _new_target_plan() -> dict:
    return {
        "mode": "new",
        "name": "Import Juli 2026",
        "valid_from": "2026-07-01",
        "valid_to": "2026-07-31",
    }


def _all_skip(keys) -> dict[str, dict]:
    return {k: {"action": "skip"} for k in keys}


def _commit(db, file_bytes: bytes, resolutions_dict: dict):
    resolutions = CommitResolutions.model_validate(resolutions_dict)
    return import_commit_service.commit_import(db, file_bytes, resolutions)


# ---------------------------------------------------------------------------
# Departments
# ---------------------------------------------------------------------------


def test_commit_creates_new_department(db, file_bytes, parsed):
    raw_depts = {r.raw_department for r in parsed.rows}
    raw_doctors = {r.raw_name for r in parsed.rows}
    target_dept = "ITS"
    assert target_dept in raw_depts

    dept_resolutions = _all_skip(raw_depts)
    dept_resolutions[target_dept] = {"action": "create"}

    result = _commit(
        db,
        file_bytes,
        {
            "target_plan": _new_target_plan(),
            "department_resolutions": dept_resolutions,
            "doctor_resolutions": _all_skip(raw_doctors),
            "code_resolutions": {},
        },
    )

    assert result.created_departments == 1
    created = dept_repo.get_department_by_name(db, target_dept)
    assert created is not None
    assert created.name == target_dept


def test_commit_maps_existing_department(db, file_bytes, parsed):
    raw_depts = {r.raw_department for r in parsed.rows}
    raw_doctors = {r.raw_name for r in parsed.rows}
    target_dept = "ITS"

    existing = dept_repo.create_department(db, {"name": "Vorhandene ITS"})
    db.commit()
    existing_count_before = db.query(Department).count()

    dept_resolutions = _all_skip(raw_depts)
    dept_resolutions[target_dept] = {"action": "map", "id": existing.id}

    result = _commit(
        db,
        file_bytes,
        {
            "target_plan": _new_target_plan(),
            "department_resolutions": dept_resolutions,
            "doctor_resolutions": _all_skip(raw_doctors),
            "code_resolutions": {},
        },
    )

    assert result.created_departments == 0
    # Keine neuen Departments angelegt.
    assert db.query(Department).count() == existing_count_before


# ---------------------------------------------------------------------------
# Doctors + EmploymentPeriods
# ---------------------------------------------------------------------------


def test_commit_creates_new_doctor_with_ep(db, file_bytes, parsed):
    raw_depts = {r.raw_department for r in parsed.rows}
    raw_doctors = {r.raw_name for r in parsed.rows}
    # Arzt mit Prozentangabe im Rohnamen wählen, z. B. 'Balck, Alexander (50%)'.
    target_doctor = next(n for n in raw_doctors if "(" in n)

    doctor_resolutions = _all_skip(raw_doctors)
    doctor_resolutions[target_doctor] = {"action": "create", "percentage": 70}

    result = _commit(
        db,
        file_bytes,
        {
            "target_plan": _new_target_plan(),
            "department_resolutions": _all_skip(raw_depts),
            "doctor_resolutions": doctor_resolutions,
            "code_resolutions": {},
        },
    )

    assert result.created_doctors == 1
    assert result.created_employment_periods == 1

    doctors = db.query(Doctor).all()
    assert len(doctors) == 1
    # Name ohne '(NN%)'-Suffix.
    assert "(" not in doctors[0].name
    assert "%" not in doctors[0].name

    eps = db.query(EmploymentPeriod).filter_by(doctor_id=doctors[0].id).all()
    assert len(eps) == 1
    assert eps[0].employment_percentage == 70
    assert eps[0].valid_from == date(2026, 7, 1)
    assert eps[0].valid_to is None


def test_commit_create_doctor_without_percentage_no_ep(db, file_bytes, parsed):
    raw_depts = {r.raw_department for r in parsed.rows}
    raw_doctors = {r.raw_name for r in parsed.rows}
    target_doctor = next(iter(raw_doctors))

    doctor_resolutions = _all_skip(raw_doctors)
    doctor_resolutions[target_doctor] = {"action": "create"}

    result = _commit(
        db,
        file_bytes,
        {
            "target_plan": _new_target_plan(),
            "department_resolutions": _all_skip(raw_depts),
            "doctor_resolutions": doctor_resolutions,
            "code_resolutions": {},
        },
    )

    assert result.created_doctors == 1
    assert result.created_employment_periods == 0
    assert db.query(EmploymentPeriod).count() == 0


# ---------------------------------------------------------------------------
# Rotations
# ---------------------------------------------------------------------------


def test_commit_creates_rotations(db, file_bytes, parsed):
    raw_depts = {r.raw_department for r in parsed.rows}
    raw_doctors = {r.raw_name for r in parsed.rows}
    # Genau eine Zeile vollständig auflösen (dept create + doctor create).
    row = parsed.rows[0]

    dept_resolutions = _all_skip(raw_depts)
    dept_resolutions[row.raw_department] = {"action": "create"}
    doctor_resolutions = _all_skip(raw_doctors)
    doctor_resolutions[row.raw_name] = {"action": "create"}

    result = _commit(
        db,
        file_bytes,
        {
            "target_plan": _new_target_plan(),
            "department_resolutions": dept_resolutions,
            "doctor_resolutions": doctor_resolutions,
            "code_resolutions": {},
        },
    )

    assert result.created_rotations == 1
    rotations = db.query(RotationAssignment).filter_by(plan_id=result.plan_id).all()
    assert len(rotations) == 1
    ra = rotations[0]
    dept = dept_repo.get_department_by_name(db, row.raw_department)
    assert ra.department_id == dept.id
    assert ra.valid_from == date(2026, 7, 1)
    assert ra.valid_to == date(2026, 7, 31)
    assert ra.is_einarbeitung is False


def test_commit_deduplicates_rotations(db, file_bytes):
    """Zweimal dasselbe (doctor, dept)-Paar → nur eine Rotation."""
    # Synthetischer Plan mit doppelter Zeile statt Fixture, um Determinismus
    # für das Dedup-Verhalten zu garantieren.
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Juli 2026"
    ws["A2"] = "Bereich"
    ws["B2"] = "Assistent"
    ws.cell(row=2, column=3, value=1)
    # Zwei identische Roster-Zeilen (gleicher Bereich, gleicher Arzt).
    ws.cell(row=3, column=1, value="ITS")
    ws.cell(row=3, column=2, value="Mustermann, Max")
    ws.cell(row=4, column=1, value="ITS")
    ws.cell(row=4, column=2, value="Mustermann, Max")
    # Dritte Zeile, damit ws.max_row > 2 sicher gilt.
    ws.cell(row=5, column=1, value="SU")
    ws.cell(row=5, column=2, value="Mustermann, Max")
    buf = BytesIO()
    wb.save(buf)
    synthetic = buf.getvalue()

    resolutions = {
        "target_plan": _new_target_plan(),
        "department_resolutions": {
            "ITS": {"action": "create"},
            "SU": {"action": "create"},
        },
        "doctor_resolutions": {
            "Mustermann, Max": {"action": "create"},
        },
        "code_resolutions": {},
    }

    result = _commit(db, synthetic, resolutions)

    # (Max, ITS) erscheint zweimal → 1 Rotation; (Max, SU) → 1 weitere = 2.
    assert result.created_rotations == 2
    rotations = db.query(RotationAssignment).filter_by(plan_id=result.plan_id).all()
    pairs = {(r.doctor_id, r.department_id) for r in rotations}
    assert len(pairs) == 2
    assert len(rotations) == 2


def test_commit_skipped_entities_excluded(db, file_bytes, parsed):
    """Skip → keine Entität, keine Rotation für übersprungene Zeilen."""
    raw_depts = {r.raw_department for r in parsed.rows}
    raw_doctors = {r.raw_name for r in parsed.rows}
    row = parsed.rows[0]

    # Bereich create, Arzt skip → Zeile kann nicht aufgelöst werden.
    dept_resolutions = _all_skip(raw_depts)
    dept_resolutions[row.raw_department] = {"action": "create"}
    doctor_resolutions = _all_skip(raw_doctors)  # alle Ärzte geskippt

    result = _commit(
        db,
        file_bytes,
        {
            "target_plan": _new_target_plan(),
            "department_resolutions": dept_resolutions,
            "doctor_resolutions": doctor_resolutions,
            "code_resolutions": {},
        },
    )

    assert result.created_doctors == 0
    assert db.query(Doctor).count() == 0
    # Keine Rotationen, weil kein Arzt aufgelöst werden konnte.
    assert result.created_rotations == 0
    assert db.query(RotationAssignment).count() == 0


def test_ep_created_for_new_doctor_in_existing_plan(db: Session, file_bytes: bytes) -> None:
    """Neue Ärzte bekommen EmploymentPeriod auch bei Ziel-Plan mode='existing'."""
    from app.repositories import plan_repository as _plan_repo

    # Parse um einen echten Raw-Namen zu bekommen
    parsed_sheet = parse_besetzungsplan(file_bytes)
    raw_name = parsed_sheet.rows[0].raw_name
    raw_dept = parsed_sheet.rows[0].raw_department

    dept = Department(name=raw_dept)
    db.add(dept)
    db.flush()

    plan = _plan_repo.create_plan(
        db,
        {"name": "Test", "valid_from": date(2026, 7, 1), "valid_to": date(2026, 7, 31)},
    )
    db.commit()

    resolutions = CommitResolutions.model_validate({
        "target_plan": {"mode": "existing", "plan_id": plan.id},
        "department_resolutions": {raw_dept: {"action": "map", "id": dept.id}},
        "doctor_resolutions": {raw_name: {"action": "create", "percentage": 75}},
        "code_resolutions": {},
    })

    import_commit_service.commit_import(db, file_bytes, resolutions)

    eps = db.query(EmploymentPeriod).all()
    assert len(eps) == 1
    assert eps[0].employment_percentage == 75
    assert eps[0].valid_from == date(2026, 7, 1)
