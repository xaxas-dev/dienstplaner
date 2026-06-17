"""Tests für den Import-Commit-Service Phase D (Tages-Codes → Absences + Shifts)."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook

from app.models.absence import Absence, AbsenceType
from app.models.shift import Shift
from app.models.shift_type import ShiftType
from app.schemas.excel_import import CommitResolutions
from app.services import import_commit_service


def _new_target_plan() -> dict:
    return {
        "mode": "new",
        "name": "Import Juli 2026",
        "valid_from": "2026-07-01",
        "valid_to": "2026-07-31",
    }


def _commit(db, file_bytes: bytes, resolutions_dict: dict):
    resolutions = CommitResolutions.model_validate(resolutions_dict)
    return import_commit_service.commit_import(db, file_bytes, resolutions)


def _make_sheet(rows: list[tuple[str, str, dict[int, str]]], day_cols: list[int]) -> bytes:
    """Baut ein synthetisches Besetzungsplan-Sheet (Juli 2026).

    rows: Liste von (bereich, arzt, {tag: code}).
    day_cols: Tag-Nummern, die als Spalten-Header (Zeile 2) erscheinen.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli 2026"
    ws["A2"] = "Bereich"
    ws["B2"] = "Assistent"
    # Tag-Spalten ab Spalte 3.
    col_of_day: dict[int, int] = {}
    for idx, day in enumerate(day_cols):
        col = 3 + idx
        ws.cell(row=2, column=col, value=day)
        col_of_day[day] = col

    for r_idx, (dept, name, cells) in enumerate(rows):
        excel_row = 3 + r_idx
        ws.cell(row=excel_row, column=1, value=dept)
        ws.cell(row=excel_row, column=2, value=name)
        for day, code in cells.items():
            ws.cell(row=excel_row, column=col_of_day[day], value=code)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _resolutions(code_resolutions: dict) -> dict:
    return {
        "target_plan": _new_target_plan(),
        "department_resolutions": {"ITS": {"action": "create"}},
        "doctor_resolutions": {},  # gefüllt vom Aufrufer
        "code_resolutions": code_resolutions,
    }


def test_absence_range_created_for_U_code(db):
    """U-Code an 5 konsekutiven Tagen → 1 Absence (URLAUB, 5-Tage-Range)."""
    cells = {d: "U" for d in (1, 2, 3, 4, 5)}
    file_bytes = _make_sheet([("ITS", "Mustermann, Max", cells)], list(range(1, 8)))

    res = _resolutions({"U": {"action": "absence", "absence_type": "URLAUB"}})
    res["doctor_resolutions"] = {"Mustermann, Max": {"action": "create"}}

    result = _commit(db, file_bytes, res)

    assert result.created_absences == 1
    absences = db.query(Absence).all()
    assert len(absences) == 1
    a = absences[0]
    assert a.absence_type == AbsenceType.URLAUB
    assert a.valid_from == date(2026, 7, 1)
    assert a.valid_to == date(2026, 7, 5)


def test_absence_separate_ranges(db):
    """U-Code mit Lücke → 2 getrennte Absence-Datensätze."""
    cells = {1: "U", 2: "U", 3: "U", 6: "U", 7: "U"}
    file_bytes = _make_sheet([("ITS", "Mustermann, Max", cells)], list(range(1, 8)))

    res = _resolutions({"U": {"action": "absence", "absence_type": "URLAUB"}})
    res["doctor_resolutions"] = {"Mustermann, Max": {"action": "create"}}

    result = _commit(db, file_bytes, res)

    assert result.created_absences == 2
    absences = sorted(db.query(Absence).all(), key=lambda x: x.valid_from)
    assert len(absences) == 2
    assert (absences[0].valid_from, absences[0].valid_to) == (date(2026, 7, 1), date(2026, 7, 3))
    assert (absences[1].valid_from, absences[1].valid_to) == (date(2026, 7, 6), date(2026, 7, 7))


def test_shift_created_for_N_star_code(db):
    """N*-Code an 1 Tag → 1 Shift mit korrektem Datum/Typ/Arzt."""
    db.add(ShiftType(name="Nachtdienst", short_name="N*"))
    db.commit()
    st = db.query(ShiftType).filter(ShiftType.short_name == "N*").first()

    cells = {10: "N*"}
    file_bytes = _make_sheet([("ITS", "Mustermann, Max", cells)], list(range(1, 31)))

    res = _resolutions({"N*": {"action": "shift", "shift_type_id": st.id}})
    res["doctor_resolutions"] = {"Mustermann, Max": {"action": "create"}}

    result = _commit(db, file_bytes, res)

    assert result.created_shifts == 1
    assigned = [s for s in db.query(Shift).all() if s.doctor_id is not None]
    assert len(assigned) == 1
    s = assigned[0]
    assert s.shift_date == date(2026, 7, 10)
    assert s.shift_type_id == st.id


def test_shift_collision_warning(db):
    """2 Ärzte mit N* am selben Tag → 1 Shift + 1 Warnung."""
    db.add(ShiftType(name="Nachtdienst", short_name="N*"))
    db.commit()
    st = db.query(ShiftType).filter(ShiftType.short_name == "N*").first()

    file_bytes = _make_sheet(
        [
            ("ITS", "Mustermann, Max", {10: "N*"}),
            ("ITS", "Beispiel, Bea", {10: "N*"}),
        ],
        list(range(1, 31)),
    )

    res = _resolutions({"N*": {"action": "shift", "shift_type_id": st.id}})
    res["doctor_resolutions"] = {
        "Mustermann, Max": {"action": "create"},
        "Beispiel, Bea": {"action": "create"},
    }

    result = _commit(db, file_bytes, res)

    assert result.created_shifts == 1
    assert db.query(Shift).filter(Shift.doctor_id.isnot(None)).count() == 1
    collision_warnings = [w for w in result.warnings if "Kollision" in w]
    assert len(collision_warnings) == 1


def test_create_shift_type_on_commit(db):
    """CodeResolutionCreateShift → neuer ShiftType + Shift damit."""
    cells = {12: "X1"}
    file_bytes = _make_sheet([("ITS", "Mustermann, Max", cells)], list(range(1, 31)))

    res = _resolutions(
        {"X1": {"action": "create_shift", "short_name": "X1", "name": "Sonderdienst X1"}}
    )
    res["doctor_resolutions"] = {"Mustermann, Max": {"action": "create"}}

    result = _commit(db, file_bytes, res)

    assert result.created_shifts == 1
    st = db.query(ShiftType).filter(ShiftType.short_name == "X1").first()
    assert st is not None
    assert st.name == "Sonderdienst X1"

    assigned = [s for s in db.query(Shift).all() if s.doctor_id is not None]
    assert len(assigned) == 1
    assert assigned[0].shift_type_id == st.id
    assert assigned[0].shift_date == date(2026, 7, 12)


def test_ignore_code_creates_nothing(db):
    """Code mit action='ignore' → keine Absence, keine Shift."""
    cells = {3: "ZZ", 4: "ZZ"}
    file_bytes = _make_sheet([("ITS", "Mustermann, Max", cells)], list(range(1, 31)))

    res = _resolutions({"ZZ": {"action": "ignore"}})
    res["doctor_resolutions"] = {"Mustermann, Max": {"action": "create"}}

    result = _commit(db, file_bytes, res)

    assert result.created_absences == 0
    assert result.created_shifts == 0
    assert db.query(Absence).count() == 0
    assert db.query(Shift).count() == 0
