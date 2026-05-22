"""Tests für GET /api/plans/{plan_id}/export."""
from datetime import date
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.plan import Plan, PlanStatus


def _make_plan(db: Session, *, name: str = "Export-Testplan") -> Plan:
    p = Plan(
        name=name,
        valid_from=date(2026, 6, 1),
        valid_to=date(2026, 6, 30),
        status=PlanStatus.DRAFT,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return p


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_export_returns_200_xlsx_mime(client: TestClient, db: Session) -> None:
    plan = _make_plan(db)
    resp = client.get(f"/api/plans/{plan.id}/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == _XLSX_MIME


def test_export_content_disposition_contains_slug_filename(
    client: TestClient, db: Session
) -> None:
    plan = _make_plan(db, name="Juni 2026")
    resp = client.get(f"/api/plans/{plan.id}/export")
    assert resp.status_code == 200
    cd = resp.headers.get("content-disposition", "")
    assert "attachment" in cd
    assert "Juni-2026.xlsx" in cd


def test_export_unknown_plan_returns_404(client: TestClient) -> None:
    resp = client.get("/api/plans/999999/export")
    assert resp.status_code == 404


def test_export_response_body_opens_as_workbook(client: TestClient, db: Session) -> None:
    plan = _make_plan(db)
    resp = client.get(f"/api/plans/{plan.id}/export")
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    assert "Dienste" in wb.sheetnames
    ws = wb["Dienste"]
    header = [cell.value for cell in ws[1]]
    assert "Datum" in header
    assert "Arzt" in header
