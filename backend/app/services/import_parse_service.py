"""Reine Parse-Logik für den Besetzungsplan-Excel (kein DB-Zugriff)."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import load_workbook

# Deutsche Monatsnamen → Monatszahl. Lange und kurze Schreibweisen.
_MONTH_MAP: dict[str, int] = {
    "januar": 1,
    "jan": 1,
    "februar": 2,
    "feb": 2,
    "märz": 3,
    "maerz": 3,
    "mär": 3,
    "mrz": 3,
    "april": 4,
    "apr": 4,
    "mai": 5,
    "juni": 6,
    "jun": 6,
    "juli": 7,
    "jul": 7,
    "august": 8,
    "aug": 8,
    "september": 9,
    "sep": 9,
    "oktober": 10,
    "okt": 10,
    "november": 11,
    "nov": 11,
    "dezember": 12,
    "dez": 12,
}

_TITLE_RE = re.compile(r"(\w+)\s+(\d{4})")
_PLACEHOLDER = "XXXX"


class ParseError(ValueError):
    """Der Besetzungsplan konnte nicht geparst werden."""


@dataclass
class ParsedRow:
    raw_department: str
    raw_name: str
    cells: dict[int, str]  # day_number (1-31) → code string


@dataclass
class ParsedSheet:
    sheet_name: str
    year: int
    month: int
    rows: list[ParsedRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _parse_title(title: str) -> tuple[int, int]:
    """Parst 'Juli 2026' → (month=7, year=2026)."""
    match = _TITLE_RE.search(title)
    if match is None:
        raise ParseError(f"Sheet-Titel '{title}' enthält kein 'Monat JJJJ'-Muster.")
    month_raw, year_raw = match.group(1), match.group(2)
    month = _MONTH_MAP.get(month_raw.strip().lower())
    if month is None:
        raise ParseError(f"Unbekannter Monatsname '{month_raw}' im Sheet-Titel '{title}'.")
    return month, int(year_raw)


def parse_besetzungsplan(file_bytes: bytes) -> ParsedSheet:
    """Parst eine Besetzungsplan-Excel-Datei zu einer ParsedSheet.

    Nutzt das erste nicht-leere Blatt. Zeile 2 definiert die Tag-Spalten,
    Zeilen ab 3 sind Roster-Zeilen (Col A + Col B gesetzt).
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    ws = None
    for sheet in wb.worksheets:
        if sheet.max_row and sheet.max_row > 2:
            ws = sheet
            break
    if ws is None:
        raise ParseError("Keine nicht-leere Tabelle in der Datei gefunden.")

    month, year = _parse_title(str(ws.title))
    parsed = ParsedSheet(sheet_name=str(ws.title), year=year, month=month)

    # Zeile 2: Spalte → Tag-Nummer-Mapping aufbauen (nur ganzzahlige Werte).
    col_to_day: dict[int, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=2, column=col).value
        if isinstance(value, int):
            col_to_day[col] = value
        elif isinstance(value, str) and value.strip().isdigit():
            col_to_day[col] = int(value.strip())

    consecutive_empty = 0
    for row in range(3, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1).value
        cell_b = ws.cell(row=row, column=2).value

        if cell_a is None and cell_b is None:
            consecutive_empty += 1
            if consecutive_empty >= 3:
                break
            continue
        consecutive_empty = 0

        # Nur echte Roster-Zeilen: Bereich UND Assistent gesetzt.
        if cell_a is None or cell_b is None:
            continue

        raw_department = str(cell_a).strip()
        raw_name = str(cell_b).strip()

        if raw_name == _PLACEHOLDER:
            parsed.warnings.append(
                f"Zeile {row}: Platzhalter '{_PLACEHOLDER}' in Spalte Assistent übersprungen "
                f"(Bereich '{raw_department}')."
            )
            continue

        cells: dict[int, str] = {}
        for col, day in col_to_day.items():
            code = ws.cell(row=row, column=col).value
            if code is None:
                continue
            code_str = str(code).strip()
            if code_str:
                cells[day] = code_str

        parsed.rows.append(
            ParsedRow(raw_department=raw_department, raw_name=raw_name, cells=cells)
        )

    return parsed
