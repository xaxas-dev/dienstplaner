"""Excel-Export für einen Plan (read-only, openpyxl, in-memory BytesIO)."""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.repositories import plan_repository
from app.services.exceptions import PlanNotFoundError

_WEEKDAYS_DE = ("Mo", "Di", "Mi", "Do", "Fr", "Sa", "So")
_HEADER = [
    "Datum",
    "Wochentag",
    "Schichttyp (Kurz)",
    "Schichttyp",
    "Arzt-Kürzel",
    "Arzt",
    "Gepinnt",
    "Notiz",
]


def make_filename_slug(plan_name: str, plan_id: int) -> str:
    slug = re.sub(r"[^A-Za-z0-9_-]+", "-", plan_name).strip("-")
    return f"{slug}.xlsx" if slug else f"plan-{plan_id}.xlsx"


def build_plan_xlsx(db: Session, plan_id: int) -> bytes:
    """Baut ein Excel-Workbook für plan_id und gibt die Bytes zurück.

    Raises:
        PlanNotFoundError: plan_id existiert nicht.
    """
    plan = plan_repository.get_plan(db, plan_id)
    if plan is None:
        raise PlanNotFoundError(plan_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dienste"

    # Header-Zeile (fett)
    ws.append(_HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Shifts sortieren: shift_date ASC, dann display_order ASC
    shifts = sorted(
        plan.shifts,
        key=lambda s: (s.shift_date, s.shift_type.display_order if s.shift_type else 0),
    )

    col_widths = [len(h) for h in _HEADER]

    for shift in shifts:
        datum = shift.shift_date.isoformat()
        wochentag = _WEEKDAYS_DE[shift.shift_date.weekday()]
        st_kurz = shift.shift_type.short_name if shift.shift_type else ""
        st_name = shift.shift_type.name if shift.shift_type else ""
        arzt_kurz = shift.doctor.short_name if shift.doctor else ""
        arzt_name = shift.doctor.name if shift.doctor else ""
        gepinnt = "ja" if shift.is_pinned else ""
        notiz = shift.notes or ""

        row = [datum, wochentag, st_kurz, st_name, arzt_kurz, arzt_name, gepinnt, notiz]
        ws.append(row)

        for i, val in enumerate(row):
            col_widths[i] = max(col_widths[i], len(str(val)))

    # Rudimentäres Spalten-Autofit (geklemmmt auf [8, 40])
    for i, width in enumerate(col_widths):
        col_letter = ws.cell(row=1, column=i + 1).column_letter
        ws.column_dimensions[col_letter].width = max(8, min(40, width + 2))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
